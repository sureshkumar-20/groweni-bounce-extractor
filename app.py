"""
Groweni Bounce Email Extractor v1.0 - Streamlit Version
"""

import streamlit as st
import re
import csv
import zipfile
import io
import os
from datetime import datetime
from pathlib import Path

try:
    import extract_msg
    MSG_SUPPORT = True
except ImportError:
    MSG_SUPPORT = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# ─────────────────────────────────────────────
# BOUNCE PATTERNS
# ─────────────────────────────────────────────

BOUNCE_PATTERNS = [
    r"Your message to\s+([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})\s+has been blocked",
    r"Your message to\s+([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})\s+could not be delivered",
    r"Delivery to the following recipient[s]? failed[^:]*:\s*([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    r"delivery to\s+([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})\s+failed",
    r"Final-Recipient:\s*rfc822;\s*([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    r"Original-Recipient:\s*rfc822;\s*([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    r"Recipient Address:\s*([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    r"The following recipient\(s\) cannot be reached:\s*['\"]?([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    r"550[- ].*?([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    r"5\.1\.1.*?([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
]

BOUNCE_KEYWORDS = [
    "undeliverable", "delivery status notification", "mail delivery failed",
    "delivery failure", "returned mail", "message blocked", "message rejected",
    "failed delivery", "could not be delivered", "bounce", "non-delivery"
]

def is_bounce(text):
    t = text.lower()
    return any(k in t for k in BOUNCE_KEYWORDS)

def extract_emails(text):
    found = []
    for pattern in BOUNCE_PATTERNS:
        found.extend(re.findall(pattern, text, re.IGNORECASE | re.MULTILINE))
    if is_bounce(text):
        found.extend(re.findall(r"[\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,}", text))
    return [e.lower().strip() for e in found]

def process_csv(file_bytes):
    emails = []
    logs = []
    try:
        content = file_bytes.decode("utf-8", errors="ignore")
        reader = csv.DictReader(io.StringIO(content))
        for i, row in enumerate(reader):
            text = " ".join(str(v) for v in row.values())
            if is_bounce(text):
                found = extract_emails(text)
                emails.extend(found)
                if found:
                    logs.append(f"Row {i+2}: {found}")
    except Exception as e:
        logs.append(f"CSV error: {e}")
    return emails, logs

def process_msg_bytes(file_bytes, filename):
    emails = []
    logs = []
    if not MSG_SUPPORT:
        logs.append("⚠ extract_msg not available")
        return emails, logs
    try:
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".msg", delete=False) as tmp:
            tmp.write(file_bytes)
            tmp_path = tmp.name
        msg = extract_msg.Message(tmp_path)
        body = msg.body or ""
        subject = msg.subject or ""
        os.unlink(tmp_path)
        if is_bounce(subject) or is_bounce(body):
            found = extract_emails(body)
            emails.extend(found)
            logs.append(f"{filename}: {found}")
    except Exception as e:
        logs.append(f"MSG error {filename}: {e}")
    return emails, logs

def process_zip(file_bytes):
    emails = []
    logs = []
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes)) as z:
            msg_files = [f for f in z.namelist() if f.lower().endswith(".msg")]
            logs.append(f"ZIP: {len(msg_files)} MSG files found")
            for mf in msg_files:
                data = z.read(mf)
                e, l = process_msg_bytes(data, Path(mf).name)
                emails.extend(e)
                logs.extend(l)
    except Exception as e:
        logs.append(f"ZIP error: {e}")
    return emails, logs

def to_excel(emails):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blocked Recipients"
    ws["A1"] = "Email Address"
    ws["B1"] = "Extracted On"
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["B1"].font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", start_color="1F4E78")
    ws["A1"].fill = fill
    ws["B1"].fill = fill
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    for i, email in enumerate(emails, start=2):
        ws.cell(row=i, column=1, value=email)
        ws.cell(row=i, column=2, value=ts)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def to_csv(emails):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Email Address", "Extracted On"])
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    for email in emails:
        writer.writerow([email, ts])
    return buf.getvalue().encode("utf-8")

# ─────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────

st.set_page_config(
    page_title="Groweni Bounce Extractor",
    page_icon="📧",
    layout="centered"
)

st.markdown("""
<style>
.main { background-color: #f0f4f8; }
.stButton>button {
    background-color: #1F4E78;
    color: white;
    border-radius: 8px;
    font-weight: bold;
    width: 100%;
}
.stButton>button:hover { background-color: #2980b9; }
</style>
""", unsafe_allow_html=True)

st.markdown("## 📧 Groweni Bounce Email Extractor")
st.markdown("Extract blocked/bounced emails from **CSV, MSG, ZIP** files automatically.")
st.divider()

# Upload
uploaded_files = st.file_uploader(
    "Upload Files (CSV / MSG / ZIP)",
    type=["csv", "msg", "zip"],
    accept_multiple_files=True,
    help="Select one or more files. ZIP should contain MSG files."
)

# Export format
fmt = st.radio("Export Format", ["Excel (.xlsx)", "CSV (.csv)"], horizontal=True)

st.divider()

if st.button("🚀 Extract Bounce Emails"):
    if not uploaded_files:
        st.warning("Please upload at least one file.")
    else:
        all_emails = []
        all_logs = []

        with st.spinner("Processing files..."):
            for f in uploaded_files:
                ext = Path(f.name).suffix.lower()
                file_bytes = f.read()
                st.write(f"📂 Processing: **{f.name}**")

                if ext == ".csv":
                    e, l = process_csv(file_bytes)
                elif ext == ".msg":
                    e, l = process_msg_bytes(file_bytes, f.name)
                elif ext == ".zip":
                    e, l = process_zip(file_bytes)
                else:
                    e, l = [], [f"Unknown format: {ext}"]

                all_emails.extend(e)
                all_logs.extend(l)
                st.write(f"  → {len(e)} emails found")

        # Deduplicate
        unique = sorted(set(all_emails))

        st.divider()

        if not unique:
            st.error("❌ No bounce emails found in uploaded files.")
        else:
            st.success(f"✅ **{len(unique)} unique bounce emails extracted!**")

            # Show table
            st.dataframe({"Email Address": unique}, use_container_width=True)

            # Download
            ts = datetime.now().strftime("%Y%m%d_%H%M")
            if "Excel" in fmt and EXCEL_SUPPORT:
                data = to_excel(unique)
                st.download_button(
                    "⬇ Download Excel",
                    data=data,
                    file_name=f"Blocked_Recipients_{ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                data = to_csv(unique)
                st.download_button(
                    "⬇ Download CSV",
                    data=data,
                    file_name=f"Blocked_Recipients_{ts}.csv",
                    mime="text/csv"
                )

        # Logs
        with st.expander("📋 Processing Log"):
            for log in all_logs:
                st.text(log)

st.divider()
st.caption("Groweni Mail Automation System v1.0 | Supports Gmail, Outlook, Yahoo, Microsoft 365 bounce formats")
