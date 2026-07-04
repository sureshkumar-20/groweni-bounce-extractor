"""
Groweni Bounce Email Extractor v1.0
Supports: CSV, MSG, ZIP(MSG files), PST
Extracts blocked/bounced recipient emails and exports to Excel/CSV
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import os
import re
import csv
import zipfile
import json
from pathlib import Path
from datetime import datetime

# Optional imports - graceful fallback
try:
    import extract_msg
    MSG_SUPPORT = True
except ImportError:
    MSG_SUPPORT = False

try:
    import pypff
    PST_SUPPORT = True
except ImportError:
    PST_SUPPORT = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_SUPPORT = True
except ImportError:
    EXCEL_SUPPORT = False

# ─────────────────────────────────────────────
# BOUNCE PATTERN ENGINE
# ─────────────────────────────────────────────

BOUNCE_PATTERNS = [
    # Gmail / Google
    r"Your message to\s+([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})\s+has been blocked",
    r"Your message to\s+([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})\s+could not be delivered",
    # Generic delivery failure
    r"Delivery to the following recipient[s]? failed[^:]*:\s*([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    r"delivery to\s+([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})\s+failed",
    # RFC/SMTP headers
    r"Final-Recipient:\s*rfc822;\s*([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    r"Original-Recipient:\s*rfc822;\s*([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    # Microsoft / Outlook NDR
    r"Recipient Address:\s*([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    r"The following recipient\(s\) cannot be reached:\s*['\"]?([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    # Yahoo
    r"address was not found in Yahoo's system:\s*([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    # 550 errors
    r"550[- ].*?([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    r"5\.1\.1.*?([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
    # Generic To: pattern in bounce bodies
    r"to:\s*([\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,})",
]

BOUNCE_SUBJECT_KEYWORDS = [
    "undeliverable", "delivery status notification", "mail delivery failed",
    "delivery failure", "returned mail", "message blocked", "message rejected",
    "delivery notification", "failed delivery", "ndr", "non-delivery",
    "could not be delivered", "bounce"
]

def is_bounce_subject(subject: str) -> bool:
    s = subject.lower()
    return any(kw in s for kw in BOUNCE_SUBJECT_KEYWORDS)

def extract_emails_from_text(text: str) -> list:
    """Extract bounced recipient emails using all patterns."""
    found = []
    for pattern in BOUNCE_PATTERNS:
        matches = re.findall(pattern, text, re.IGNORECASE | re.MULTILINE)
        found.extend(matches)
    # Fallback: any email in text if bounce indicators present
    if any(kw in text.lower() for kw in ["blocked", "rejected", "failed", "undeliverable", "bounce"]):
        generic = re.findall(r"[\w._%+\-]+@[\w.\-]+\.[a-zA-Z]{2,}", text)
        found.extend(generic)
    return [e.lower().strip() for e in found if e]

# ─────────────────────────────────────────────
# FILE PROCESSORS
# ─────────────────────────────────────────────

def process_csv(filepath: str, log_fn) -> list:
    """Process CSV file — look for bounce indicators in any text column."""
    emails = []
    try:
        with open(filepath, "r", encoding="utf-8", errors="ignore") as f:
            reader = csv.DictReader(f)
            headers = reader.fieldnames or []
            log_fn(f"  CSV headers: {headers}")
            for i, row in enumerate(reader):
                full_text = " ".join(str(v) for v in row.values())
                subject = ""
                for h in headers:
                    if "subject" in h.lower():
                        subject = str(row.get(h, ""))
                        break
                if is_bounce_subject(subject) or any(
                    kw in full_text.lower() for kw in ["blocked", "rejected", "undeliverable", "failed"]
                ):
                    extracted = extract_emails_from_text(full_text)
                    emails.extend(extracted)
                    if extracted:
                        log_fn(f"  Row {i+2}: found {extracted}")
    except Exception as e:
        log_fn(f"  CSV error: {e}")
    return emails

def process_msg_file(filepath: str, log_fn) -> list:
    """Process a single .msg file."""
    emails = []
    if not MSG_SUPPORT:
        log_fn("  ⚠ extract_msg not installed — MSG support unavailable")
        return emails
    try:
        msg = extract_msg.Message(filepath)
        subject = msg.subject or ""
        body = msg.body or ""
        if is_bounce_subject(subject) or any(
            kw in body.lower() for kw in ["blocked", "rejected", "undeliverable", "failed"]
        ):
            extracted = extract_emails_from_text(body)
            emails.extend(extracted)
            log_fn(f"  MSG {Path(filepath).name}: found {extracted}")
    except Exception as e:
        log_fn(f"  MSG error {Path(filepath).name}: {e}")
    return emails

def process_zip(filepath: str, log_fn) -> list:
    """Extract and process MSG files inside a ZIP."""
    emails = []
    if not MSG_SUPPORT:
        log_fn("  ⚠ extract_msg not installed — MSG support unavailable")
        return emails
    try:
        import tempfile
        with zipfile.ZipFile(filepath, "r") as z:
            msg_files = [f for f in z.namelist() if f.lower().endswith(".msg")]
            log_fn(f"  ZIP contains {len(msg_files)} MSG files")
            with tempfile.TemporaryDirectory() as tmpdir:
                for mf in msg_files:
                    z.extract(mf, tmpdir)
                    full_path = os.path.join(tmpdir, mf)
                    emails.extend(process_msg_file(full_path, log_fn))
    except Exception as e:
        log_fn(f"  ZIP error: {e}")
    return emails

def process_pst(filepath: str, log_fn) -> list:
    """Process PST file using pypff."""
    emails = []
    if not PST_SUPPORT:
        log_fn("  ⚠ pypff not installed — PST support unavailable")
        log_fn("    Install: pip install pypff-python")
        return emails
    try:
        pst = pypff.file()
        pst.open(filepath)
        root = pst.get_root_folder()
        emails.extend(_scan_pst_folder(root, log_fn))
        pst.close()
    except Exception as e:
        log_fn(f"  PST error: {e}")
    return emails

def _scan_pst_folder(folder, log_fn) -> list:
    emails = []
    try:
        for i in range(folder.number_of_sub_messages):
            msg = folder.get_sub_message(i)
            subject = msg.subject or ""
            body = msg.plain_text_body or ""
            if isinstance(body, bytes):
                body = body.decode("utf-8", errors="ignore")
            if is_bounce_subject(subject) or any(
                kw in body.lower() for kw in ["blocked", "rejected", "undeliverable", "failed"]
            ):
                extracted = extract_emails_from_text(body)
                emails.extend(extracted)
                if extracted:
                    log_fn(f"  PST msg: found {extracted}")
        for i in range(folder.number_of_sub_folders):
            sub = folder.get_sub_folder(i)
            emails.extend(_scan_pst_folder(sub, log_fn))
    except Exception:
        pass
    return emails

# ─────────────────────────────────────────────
# EXPORT
# ─────────────────────────────────────────────

def export_excel(emails: list, output_path: str):
    if not EXCEL_SUPPORT:
        raise RuntimeError("openpyxl not installed")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blocked Recipients"
    # Header
    ws["A1"] = "Email Address"
    ws["B1"] = "Extracted On"
    ws["A1"].font = Font(bold=True, color="FFFFFF")
    ws["B1"].font = Font(bold=True, color="FFFFFF")
    fill = PatternFill("solid", start_color="1F4E78")
    ws["A1"].fill = fill
    ws["B1"].fill = fill
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 20
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    for i, email in enumerate(emails, start=2):
        ws.cell(row=i, column=1, value=email)
        ws.cell(row=i, column=2, value=ts)
    wb.save(output_path)

def export_csv(emails: list, output_path: str):
    ts = datetime.now().strftime("%Y-%m-%d %H:%M")
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["Email Address", "Extracted On"])
        for email in emails:
            writer.writerow([email, ts])

# ─────────────────────────────────────────────
# GUI
# ─────────────────────────────────────────────

class BounceExtractorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Groweni Bounce Email Extractor v1.0")
        self.root.geometry("700x600")
        self.root.resizable(True, True)
        self.root.configure(bg="#1F4E78")

        self.files = []
        self.results = []
        self._build_ui()

    def _build_ui(self):
        # Title
        tk.Label(self.root, text="📧 Groweni Bounce Email Extractor",
                 font=("Arial", 16, "bold"), bg="#1F4E78", fg="white").pack(pady=(20, 5))
        tk.Label(self.root, text="Extract blocked/bounced emails from CSV, MSG, ZIP, PST files",
                 font=("Arial", 10), bg="#1F4E78", fg="#a8d0f0").pack(pady=(0, 15))

        # Main frame
        main = tk.Frame(self.root, bg="#f0f4f8", bd=0)
        main.pack(fill="both", expand=True, padx=20, pady=(0, 20))

        # File list frame
        file_frame = tk.LabelFrame(main, text="Selected Files", font=("Arial", 10, "bold"),
                                    bg="#f0f4f8", fg="#1F4E78", padx=10, pady=10)
        file_frame.pack(fill="x", padx=15, pady=(15, 5))

        self.file_listbox = tk.Listbox(file_frame, height=5, font=("Arial", 9),
                                        selectmode=tk.MULTIPLE, bg="white", relief="flat",
                                        highlightthickness=1, highlightcolor="#1F4E78")
        self.file_listbox.pack(fill="x", side="left", expand=True)
        sb = tk.Scrollbar(file_frame, orient="vertical", command=self.file_listbox.yview)
        sb.pack(side="right", fill="y")
        self.file_listbox.configure(yscrollcommand=sb.set)

        # Buttons row
        btn_frame = tk.Frame(main, bg="#f0f4f8")
        btn_frame.pack(fill="x", padx=15, pady=5)

        self._btn(btn_frame, "➕ Add Files", self._add_files, "#2196F3").pack(side="left", padx=5)
        self._btn(btn_frame, "🗑 Remove Selected", self._remove_files, "#f44336").pack(side="left", padx=5)
        self._btn(btn_frame, "❌ Clear All", self._clear_files, "#9E9E9E").pack(side="left", padx=5)

        # Support status
        status_frame = tk.Frame(main, bg="#f0f4f8")
        status_frame.pack(fill="x", padx=15, pady=2)
        tk.Label(status_frame, text=f"MSG support: {'✅' if MSG_SUPPORT else '❌ (pip install extract-msg)'}  |  "
                                     f"PST support: {'✅' if PST_SUPPORT else '❌ (pip install pypff-python)'}  |  "
                                     f"Excel export: {'✅' if EXCEL_SUPPORT else '❌ (pip install openpyxl)'}",
                 font=("Arial", 8), bg="#f0f4f8", fg="#555").pack(anchor="w")

        # Output format
        fmt_frame = tk.LabelFrame(main, text="Export Format", font=("Arial", 10, "bold"),
                                   bg="#f0f4f8", fg="#1F4E78", padx=10, pady=8)
        fmt_frame.pack(fill="x", padx=15, pady=5)
        self.export_format = tk.StringVar(value="excel")
        tk.Radiobutton(fmt_frame, text="Excel (.xlsx)", variable=self.export_format,
                       value="excel", bg="#f0f4f8", font=("Arial", 10)).pack(side="left", padx=10)
        tk.Radiobutton(fmt_frame, text="CSV (.csv)", variable=self.export_format,
                       value="csv", bg="#f0f4f8", font=("Arial", 10)).pack(side="left", padx=10)

        # Extract button
        self._btn(main, "🚀 Extract Bounce Emails", self._run_extract, "#1F4E78",
                  font=("Arial", 12, "bold"), pady=10).pack(padx=15, pady=8, fill="x")

        # Log area
        log_frame = tk.LabelFrame(main, text="Log", font=("Arial", 10, "bold"),
                                   bg="#f0f4f8", fg="#1F4E78", padx=10, pady=5)
        log_frame.pack(fill="both", expand=True, padx=15, pady=(0, 10))
        self.log_text = tk.Text(log_frame, height=8, font=("Consolas", 8),
                                 bg="#1a1a2e", fg="#00ff88", relief="flat", state="disabled")
        self.log_text.pack(fill="both", expand=True)
        log_sb = tk.Scrollbar(log_frame, orient="vertical", command=self.log_text.yview)
        log_sb.pack(side="right", fill="y")
        self.log_text.configure(yscrollcommand=log_sb.set)

        # Result label
        self.result_label = tk.Label(main, text="", font=("Arial", 11, "bold"),
                                      bg="#f0f4f8", fg="#1F4E78")
        self.result_label.pack()

    def _btn(self, parent, text, cmd, color, font=("Arial", 10, "bold"), pady=6):
        return tk.Button(parent, text=text, command=cmd, bg=color, fg="white",
                         font=font, relief="flat", padx=12, pady=pady,
                         activebackground=color, cursor="hand2")

    def _add_files(self):
        files = filedialog.askopenfilenames(
            title="Select Files",
            filetypes=[
                ("Supported files", "*.csv *.msg *.zip *.pst"),
                ("CSV files", "*.csv"),
                ("MSG files", "*.msg"),
                ("ZIP files", "*.zip"),
                ("PST files", "*.pst"),
                ("All files", "*.*"),
            ]
        )
        for f in files:
            if f not in self.files:
                self.files.append(f)
                self.file_listbox.insert(tk.END, f"  {Path(f).name}  ({Path(f).suffix.upper()})")

    def _remove_files(self):
        selected = list(self.file_listbox.curselection())
        for i in reversed(selected):
            self.file_listbox.delete(i)
            self.files.pop(i)

    def _clear_files(self):
        self.file_listbox.delete(0, tk.END)
        self.files.clear()

    def _log(self, msg):
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state="disabled")
        self.root.update_idletasks()

    def _run_extract(self):
        if not self.files:
            messagebox.showwarning("No Files", "Please add at least one file to process.")
            return
        threading.Thread(target=self._extract_worker, daemon=True).start()

    def _extract_worker(self):
        self._log("=" * 50)
        self._log(f"Started: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        all_emails = []

        for filepath in self.files:
            ext = Path(filepath).suffix.lower()
            self._log(f"\n📂 Processing: {Path(filepath).name}")
            if ext == ".csv":
                found = process_csv(filepath, self._log)
            elif ext == ".msg":
                found = process_msg_file(filepath, self._log)
            elif ext == ".zip":
                found = process_zip(filepath, self._log)
            elif ext == ".pst":
                found = process_pst(filepath, self._log)
            else:
                self._log(f"  ⚠ Unknown format: {ext}")
                found = []
            self._log(f"  → {len(found)} emails found")
            all_emails.extend(found)

        # Deduplicate
        unique = sorted(set(all_emails))
        self._log(f"\n✅ Total unique emails: {len(unique)}")
        self.results = unique

        if not unique:
            self._log("⚠ No bounce emails found.")
            self.result_label.configure(text="No bounce emails found.", fg="red")
            return

        # Ask save location
        fmt = self.export_format.get()
        ext = ".xlsx" if fmt == "excel" else ".csv"
        default_name = f"Blocked_Recipients_{datetime.now().strftime('%Y%m%d_%H%M')}{ext}"

        save_path = filedialog.asksaveasfilename(
            title="Save Output File",
            defaultextension=ext,
            initialfile=default_name,
            filetypes=[("Excel", "*.xlsx")] if fmt == "excel" else [("CSV", "*.csv")]
        )

        if not save_path:
            self._log("Save cancelled.")
            return

        try:
            if fmt == "excel":
                export_excel(unique, save_path)
            else:
                export_csv(unique, save_path)
            self._log(f"💾 Saved: {save_path}")
            self.result_label.configure(
                text=f"✅ {len(unique)} unique emails extracted → {Path(save_path).name}", fg="#1F4E78")
            messagebox.showinfo("Done!", f"{len(unique)} unique bounce emails extracted!\n\nSaved to:\n{save_path}")
        except Exception as e:
            self._log(f"❌ Export error: {e}")
            messagebox.showerror("Export Error", str(e))

# ─────────────────────────────────────────────
# ENTRY POINT
# ─────────────────────────────────────────────

if __name__ == "__main__":
    root = tk.Tk()
    app = BounceExtractorApp(root)
    root.mainloop()
